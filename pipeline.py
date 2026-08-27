# -*- coding: utf-8 -*-
"""
SKU&SPU 清洗流水线（v9 口径，可复用模块）
========================================
把 clean_sku.py(阶段一) + build_final.py(阶段二) 合并为单一函数 run_pipeline，
供 Flask 系统调用：上传原始「SKU平台信息套件成员货品体现」CSV -> 返回清洗后数据。

输出列顺序（系统面板/下载报表展示顺序）：
  ID、上架SKU、ASIN、库存SKU、成员货品数量、listing后台状态（领星+后台）、店铺、
  唯一SKU、SPU、SPU带尺寸、包裹类型、发货SKU、是否同步平台库存、美国生命周期状态、非活动、leadtime、
  是否重复维护、异常确认状态、异常确认备注、缺失平台信息异常、
  内部 ID、德国/日本/英国生命周期状态、类型、代运营
  注：缺失平台信息异常 = 领星/walmart 数据源有、平台信息表无 的 SKU，追加为末行并标「是」。
"""
import csv
import re
import zipfile
from collections import defaultdict, Counter, OrderedDict

# ---------- 自定义排序字典（阶段一 & 最终排序共用） ----------
INACTIVE_ORDER = {'否': 0, '是': 1}          # 非活动：否 -> 是
STORE_ORDER = {
    'US-Amazon-HLLdeco-KH': 0,
    'US-Amazon-Home Nest-KH': 1,
    'US-Amazon-Urban-KH': 2,
    'US-Amazon-JYT-KH': 3,
    'US-Amazon-XYT-KH': 4,
    'US-Walmart': 5,
}
SYNC_ORDER = {'是': 0, '否': 1, '': 2}        # 是否同步：是 -> 否 -> 空
PKG_ORDER = {'单件': 0, '套件': 1, 'AB件': 2, '套组': 3, '异常': 4}
# listing后台状态（领星+后台）排序：在售 -> PUBLISHED -> 停售 -> UNPUBLISHED -> 已删除 -> SYSTEM_PROBLEM -> IN_PROGRESS
LISTING_ORDER = {
    '在售': 0,
    'PUBLISHED': 1,
    '停售': 2,
    'UNPUBLISHED': 3,
    '已删除': 4,
    'SYSTEM_PROBLEM': 5,
    'IN_PROGRESS': 6,
}

REQUIRED_COLS = ['店铺', '14K货号', '成员货品', '成员货品数量', '非活动',
                 '是否同步平台库存', 'leadtime']

# 领星清单店铺名 -> 平台表店铺名（用于 ASIN/listing 状态匹配）
STORE_MAP = {
    'HulalahomeDeco-US': 'US-Amazon-HLLdeco-KH',
    'HOME NEST-US': 'US-Amazon-Home Nest-KH',
    'Urban Living-US': 'US-Amazon-Urban-KH',
    'JYT HOME STORE-US': 'US-Amazon-JYT-KH',
    'XYT HOME-US': 'US-Amazon-XYT-KH',
}
# 平台表 Amazon 店铺走领星清单匹配；其余店铺（US-Walmart）走 walmart 后台报表
AMAZON_STORES = set(STORE_MAP.values())

# 「是否重复维护」三态值
DUP_NO = '否'        # 原「正常」
DUP_YES = '是'       # 原「重复维护，需清洗」
DUP_INACTIVE = '非活动'   # 非活动=是 的行
TAG_ORDER = {DUP_NO: 0, DUP_YES: 1, DUP_INACTIVE: 2}


def parse_leadtime(s):
    s = s.strip()
    if s == '':
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ========================== 阶段一：三步清洗 + 打标 ==========================
def _stage1(raw_path):
    with open(raw_path, encoding='utf-8-sig', newline='') as f:
        r = csv.reader(f)
        header = next(r)
        rows = [row for row in r if any(c.strip() for c in row)]

    idx = {name: i for i, name in enumerate(header)}
    missing = [c for c in REQUIRED_COLS if c not in idx]
    if missing:
        raise ValueError("原始表格缺少必要列：" + "、".join(missing))

    def sortkey(row):
        lt = parse_leadtime(row[idx['leadtime']])
        lt_key = (1, 0) if lt is None else (0, -lt)
        return (
            INACTIVE_ORDER.get(row[idx['非活动']].strip(), 99),
            STORE_ORDER.get(row[idx['店铺']].strip(), 99),
            row[idx['14K货号']].strip(),
            row[idx['成员货品']].strip(),
            SYNC_ORDER.get(row[idx['是否同步平台库存']].strip(), 2),
            lt_key,
        )

    rows_sorted = sorted(rows, key=sortkey)

    def group_key(row):
        return (
            row[idx['店铺']].strip(),
            row[idx['14K货号']].strip(),
            row[idx['成员货品']].strip(),
            row[idx['成员货品数量']].strip(),
            row[idx['非活动']].strip(),
        )

    group_seen = OrderedDict()
    for row in rows_sorted:
        group_seen.setdefault(group_key(row), []).append(row)

    for row in rows_sorted:
        grp = group_seen[group_key(row)]
        row.append(DUP_NO if grp[0] is row else DUP_YES)

    # 规则：非活动=是 的行，统一改为「非活动」（覆盖 否/是）
    for row in rows_sorted:
        if row[idx['非活动']].strip() == '是':
            row[-1] = DUP_INACTIVE

    def sortkey2(row):
        lt = parse_leadtime(row[idx['leadtime']])
        lt_key = (1, 0) if lt is None else (0, -lt)
        return (
            TAG_ORDER.get(row[-1], 1),
            INACTIVE_ORDER.get(row[idx['非活动']].strip(), 99),
            STORE_ORDER.get(row[idx['店铺']].strip(), 99),
            row[idx['14K货号']].strip(),
            row[idx['成员货品']].strip(),
            SYNC_ORDER.get(row[idx['是否同步平台库存']].strip(), 2),
            lt_key,
        )

    rows_final = sorted(rows_sorted, key=sortkey2)
    header_final = header + ['是否重复维护']
    return header_final, rows_final


# ========================== 阶段二：增字段 + 转单扩展 ==========================
def _spu(s):
    s = s.strip()
    prefix = ''
    if s.startswith('US-SA-SW-'):
        prefix = 'US-SA-SW-'
        s = s[len(prefix):]
    first_seg = s.split('-')[0]                 # 第一个连字符分隔段
    mm = re.search(r'\d+[A-Za-z]*$', first_seg)  # 到最后一个「数字段+其后粘连字母」
    if not mm:
        return prefix + first_seg
    return prefix + first_seg[:mm.end()]

# 尺寸词（独立连字符段）：床品/家具尺寸。数据实测每 SKU 最多 1 个，仅出现于第 2 段或末段。
SIZE_WORDS = {'KB', 'TB', 'QB', 'K', 'Q', 'T', 'KING', 'QUEEN', 'TWIN', 'FULL', 'TXL'}


def _spu_size(s, spu):
    """SPU带尺寸：若 SKU 任意独立段为尺寸词 => SPU-尺寸词；否则 = SPU。"""
    for seg in s.split('-'):
        if seg in SIZE_WORDS:
            return spu + '-' + seg
    return spu


def load_transfer_sheets(path):
    """读转单表【全部 sheet】。每行 = 同品的多个 SKU 名称（首行视为表头跳过）。

    返回 (per_sheet, all_rows)：
      per_sheet = [{'name': sheet名, 'rows': 有效行数}, ...]
      all_rows  = [[sku, sku, ...], ...]  合并全部 sheet
    兼容点：不再写死「增长 / 自营」两个 sheet 名，新增/改名 sheet 自动生效。
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    per_sheet, all_rows = [], []
    for sn in wb.sheetnames:
        ws = wb[sn]
        data = list(ws.iter_rows(values_only=True))
        out = []
        for row in data[1:]:                      # 跳过表头
            skus = [str(c).strip() for c in row if c not in (None, '')]
            skus = [s for s in skus if s]
            if skus:
                out.append(skus)
        per_sheet.append({'name': sn, 'rows': len(out)})
        all_rows.extend(out)
    wb.close()
    return per_sheet, all_rows


def _union_find(all_rows):
    """把每行内的 SKU 并到同一组，返回 (groups, mem2rep)。"""
    parent = {}

    def find(x):
        parent.setdefault(x, x)
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    for skus in all_rows:
        for i in range(1, len(skus)):
            union(skus[0], skus[i])
    groups = defaultdict(list)
    for s in parent:
        groups[find(s)].append(s)
    mem2rep = {s: find(s) for s in parent}
    return groups, mem2rep


def summarize_transfer(path):
    """解析转单表并返回摘要（供上传后核对）。"""
    per_sheet, all_rows = load_transfer_sheets(path)
    groups, mem2rep = _union_find(all_rows)
    multi = sum(1 for g in groups.values() if len(g) > 1)
    if not all_rows:
        raise ValueError('转单表未解析到任何数据行（每行应为同品的多个 SKU 名称，首行为表头）')
    return {
        'sheets': per_sheet,
        'total_rows': len(all_rows),
        'sku_count': len(mem2rep),
        'group_count': len(groups),
        'multi_group_count': multi,
    }


COLOR_TRAIL = {'BGE', 'GRY', 'BRN', 'BLK', 'CRY', 'BLU', 'FLR', 'GRN', 'IVY', 'NAV',
               'ORG', 'RED', 'SAP', 'WTE', 'YEL', 'DRY', 'LGY'}  # 颜色尾随码，不判 AB件
PKG_SINGLE = re.compile(r'^[A-OQ-UW-Z]$')       # 单字母包裹位（不含 P、V）
PAT_END = re.compile(r'-(?:[A-OQ-UW-Z]|F[A-Z])$')


def _is_ab_member(m):
    if PAT_END.search(m):                        # 规则1：末段包裹位
        return True
    sg = m.split('-')                            # 规则2：包裹位 + 后续非颜色码
    for j, seg in enumerate(sg):
        if PKG_SINGLE.match(seg) and j < len(sg) - 1:
            nxt = sg[j + 1]
            if nxt not in COLOR_TRAIL and re.match(r'^[A-Z0-9]{1,}$', nxt):
                return True
    return False


def load_lingxing(path):
    """领星清单 xlsx（MSKU / 状态 / 店铺 / ASIN）。

    返回 (match_map, summary)：
      match_map[(平台店铺, MSKU)] = (ASIN, listing状态)
      店铺经 STORE_MAP 映射回平台表店名；未收录店铺跳过。
      同 (店铺, MSKU) 多行取第一条。
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError('领星清单为空')
    hdr = [str(x).strip() if x is not None else '' for x in rows[0]]
    need = ['MSKU', '状态', '店铺', 'ASIN']
    for c in need:
        if c not in hdr:
            raise ValueError('领星清单缺少列：「%s」' % c)
    i = {c: hdr.index(c) for c in need}
    m, n_row = {}, 0
    for r in rows[1:]:
        if not any(str(x).strip() for x in r):
            continue
        n_row += 1
        plat = STORE_MAP.get(str(r[i['店铺']]).strip())
        msku = str(r[i['MSKU']]).strip()
        if not plat or not msku:
            continue
        key = (plat, msku)
        if key not in m:   # 同 key 多行取第一条
            m[key] = (str(r[i['ASIN']]).strip(), str(r[i['状态']]).strip())
    return m, {'rows': n_row, 'matched': len(m)}


def load_walmart(path):
    """walmart 后台报表 csv（SKU / Item ID / Publish Status …）。

    返回 (match_map, summary)：
      match_map[SKU] = (ItemID文本, Publish Status)
    Item ID 强制转文本（数字易被 Excel 科学计数）；SKU 唯一，重复取第一条。
    """
    with open(path, encoding='utf-8-sig', newline='') as f:
        r = csv.reader(f)
        hdr = [x.strip() for x in next(r)]
        rows = [row for row in r if any(c.strip() for c in row)]
    need = ['SKU', 'Item ID', 'Publish Status']
    for c in need:
        if c not in hdr:
            raise ValueError('walmart 报表缺少列：「%s」' % c)
    i = {c: hdr.index(c) for c in need}
    m = {}
    for row in rows:
        sku = row[i['SKU']].strip()
        if not sku:
            continue
        if sku not in m:
            m[sku] = (str(row[i['Item ID']]).strip(),
                      str(row[i['Publish Status']]).strip())
    return m, {'rows': len(rows), 'matched': len(m)}


def _pkg_type(g, idx):
    MEM = idx['成员货品']
    QTY = idx['成员货品数量']
    # 参与判定的行 = 【全量行】（含「是」=重复维护、含「非活动」）；相同成员货品去重，只算1条。
    # 2026-08-26 用户拍板：组内重复行也正常判定包裹，3个完全重复记录 => 去重后1条 => 单件。
    seen = set()
    judged = []
    for r in g:
        mem = r[MEM].strip()
        if mem not in seen:
            seen.add(mem)
            judged.append(r)
    if not judged:
        return ''
    # 空SKU行：成员货品为空（数据源本身空白行）不参与包裹判定，留空。
    # 2026-08-26 用户拍板：上架SKU/成员货品为空的记录不应误判为单件。
    if len(judged) == 1 and judged[0][MEM].strip() == '':
        return ''
    spus = {_spu(r[MEM]) for r in judged}
    if len(spus) > 1:
        return '套组'
    if any(_is_ab_member(r[MEM]) for r in judged):
        return 'AB件'
    if len(judged) == 1:
        try:
            qf = float(judged[0][QTY])
        except ValueError:
            qf = 0
        if qf == 1:
            return '单件'
        if qf > 1:
            return '套件'
    else:
        # 同SPU + 多条记录（去重后） + 非AB件 -> 异常（需人工核查）
        return '异常'
    return ''


def _stage2(header, rows, xlsx_path, lingxing_path=None, walmart_path=None):
    idx = {name: i for i, name in enumerate(header)}
    COL_MEM, COL_STORE, COL_14K, COL_QTY = (
        '成员货品', '店铺', '14K货号', '成员货品数量')
    if 'ID' not in idx:
        raise ValueError('原始表格缺少「ID」列，无法按 ID+店铺+14K货号 分组判定包裹类型')

    # ---- 转单 union-find（读转单表全部 sheet） ----
    per_sheet, all_rows = load_transfer_sheets(xlsx_path)
    groups, mem2rep = _union_find(all_rows)

    # ---- 可选数据源：领星清单 / walmart 后台报表（获取 ASIN + listing后台状态） ----
    lingxing_map, lx_summary = None, None
    walmart_map, wm_summary = None, None
    if lingxing_path:
        lingxing_map, lx_summary = load_lingxing(lingxing_path)
    if walmart_path:
        walmart_map, wm_summary = load_walmart(walmart_path)

    def lookup_asin(row):
        """按 店铺+14K货号 匹配 ASIN 与 listing后台状态。未命中返回 ('', '')。"""
        store = row[idx[COL_STORE]].strip()
        sku = row[idx[COL_14K]].strip()
        if not sku:
            return '', ''
        if store in AMAZON_STORES:
            hit = lingxing_map.get((store, sku)) if lingxing_map else None
        elif store == 'US-Walmart':
            hit = walmart_map.get(sku) if walmart_map else None
        else:
            hit = None
        return hit if hit else ('', '')

    # ---- 包裹类型（仅基于【原始清洗数据】，扩展前计算） ----
    # 分组键：ID + 店铺 + 14K货号（同一 ID 的多个成员货品视为同一 listing 拆出的行）
    def pkg_key(row):
        return (row[idx['ID']].strip(), row[idx[COL_STORE]].strip(), row[idx[COL_14K]].strip())

    by_key = defaultdict(list)
    for row in rows:
        by_key[pkg_key(row)].append(row)
    pkg_cache = {k: _pkg_type(g, idx) for k, g in by_key.items()}

    # ---- 赋值 唯一SKU + SPU + SPU带尺寸 + 包裹类型 + 发货SKU + ASIN + listing后台状态 ----
    n_asin = 0
    for row in rows:
        uniq_sku = row[idx[COL_MEM]]
        spu = _spu(uniq_sku)
        row.append(uniq_sku)                                 # 唯一SKU
        row.append(spu)                                      # SPU
        row.append(_spu_size(uniq_sku, spu))                 # SPU带尺寸
        row.append(pkg_cache[pkg_key(row)])                  # 包裹类型
        row.append('是')                                     # 发货SKU：原始=是
        asin, listing = lookup_asin(row)
        row.append(asin)                                     # ASIN
        row.append(listing)                                  # listing后台状态（领星+后台）
        if asin:
            n_asin += 1

    # ---- 扩展记录（转单） ----
    # row 追加顺序：唯一SKU(idx+0)、SPU(idx+1)、SPU带尺寸(idx+2)、包裹类型(idx+3)、发货SKU(idx+4)、ASIN(idx+5)、listing(idx+6)
    IDX_FAHUO = len(idx) + 4          # 发货SKU 列索引（扩展记录需改成「否」）
    expanded_groups = set()
    result = []
    n_added = 0
    for row in rows:
        result.append(row)
        X = row[idx[COL_MEM]]
        if X in mem2rep:
            rep = mem2rep[X]
            if rep not in expanded_groups:
                expanded_groups.add(rep)
                for sku in groups[rep]:
                    if sku == X:
                        continue
                    copy = list(row)
                    copy[idx[COL_MEM]] = sku
                    copy[IDX_FAHUO] = '否'   # 发货SKU：转单扩展记录=否（勿用 copy[-1]，会误改 listing）
                    # 唯一SKU/包裹类型/ASIN/listing状态 随 copy 继承（= X 及模板）
                    result.append(copy)
                    n_added += 1

    # ---- 缺失平台信息异常：数据源(领星/walmart)有、平台信息表无 的 SKU ----
    # 平台信息表 SKU 集合 = 原始清洗数据的 14K货号（上架SKU）；数据源 SKU 集合从 match_map 键重建
    plat_skus = {str(r[idx[COL_14K]]).strip() for r in rows}
    lx_skus = {msku for (_, msku) in lingxing_map} if lingxing_map else set()   # 键=(平台店铺, MSKU)
    wm_skus = set(walmart_map) if walmart_map else set()
    lx_miss = lx_skus - plat_skus
    wm_miss = wm_skus - plat_skus
    union_miss = (lx_skus | wm_skus) - plat_skus

    # ---- 完整列索引（含阶段二新增列） ----
    COL_ASIN = 'ASIN'
    COL_LISTING = 'listing后台状态（领星+后台）'
    fidx = {name: i for i, name in enumerate(
        header + ['唯一SKU', 'SPU', 'SPU带尺寸', '包裹类型', '发货SKU', COL_ASIN, COL_LISTING,
                  '异常确认状态', '异常确认备注'])}

    # ---- 新增：异常确认状态 / 异常确认备注（仅异常行默认待确认） ----
    for row in result:
        if row[fidx['包裹类型']] == '异常':
            row.append('待确认')      # 异常确认状态
            row.append('')           # 异常确认备注
        else:
            row.append('')           # 异常确认状态（非异常留空）
            row.append('')           # 异常确认备注

    # ---- 最终排序（原规则在前；发货SKU、包裹类型 置后） ----

    def final_key(row):
        fahuo = row[fidx['发货SKU']]
        fahuo_k = 0 if fahuo == '是' else 1
        pkg = row[fidx['包裹类型']]
        pkg_k = PKG_ORDER.get(pkg, 99)
        tag = row[fidx['是否重复维护']]
        tag_k = TAG_ORDER.get(tag, 1)
        listing = row[fidx[COL_LISTING]]
        listing_k = LISTING_ORDER.get(listing, 99)
        lt = parse_leadtime(row[fidx['leadtime']])
        lt_k = (1, 0) if lt is None else (0, -lt)
        return (
            tag_k,
            INACTIVE_ORDER.get(row[fidx['非活动']].strip(), 99),
            STORE_ORDER.get(row[fidx['店铺']].strip(), 99),
            row[fidx['14K货号']].strip(),
            row[fidx['成员货品']].strip(),
            SYNC_ORDER.get(row[fidx['是否同步平台库存']].strip(), 2),
            lt_k,
            fahuo_k,
            pkg_k,
            listing_k,
        )

    result.sort(key=final_key)

    # ---- 列重排：系统面板/下载报表展示顺序在前，其余列在后 ----
    ORDER = [
        'ID', '14K货号', 'ASIN', '成员货品', '成员货品数量',
        'listing后台状态（领星+后台）', '店铺', '唯一SKU', 'SPU', 'SPU带尺寸', '包裹类型', '发货SKU',
        '是否同步平台库存', '美国生命周期状态', '非活动', 'leadtime',
        '是否重复维护', '异常确认状态', '异常确认备注',
        '内部 ID', '德国生命周期状态', '日本生命周期状态', '英国生命周期状态',
        '类型', '代运营'
    ]
    RENAME = {'14K货号': '上架SKU', '成员货品': '库存SKU'}
    new_header = [RENAME.get(n, n) for n in ORDER]
    result = [[r[fidx[n]] for n in ORDER] for r in result]
    fidx = {name: i for i, name in enumerate(new_header)}

    # ---- 缺失平台信息异常：数据源(领星/walmart)有、平台信息表无 的 SKU ----
    # 组装为独立列表（不混入主表），每条包含：SKU/SPU/店铺/ASIN/listing/来源
    miss_rows = []
    if union_miss:
        lx_by_sku = {}                       # SKU -> (平台店铺, ASIN, listing)，同 SKU 取第一条
        if lingxing_map:
            for (plat, sku), (asin, listing) in lingxing_map.items():
                if sku not in lx_by_sku:
                    lx_by_sku[sku] = (plat, asin, listing)
        for sku in sorted(union_miss):
            src = lx_by_sku.get(sku)
            if src:                          # 领星优先
                store, asin, listing = src
                source = 'lingxing'
            else:                            # 否则 walmart（默认店铺 US-Walmart）
                store, asin, listing = 'US-Walmart', *walmart_map[sku]
                source = 'walmart'
            miss_rows.append({
                'sku': sku,
                'spu': _spu(sku),
                'store': store,
                'asin': asin,
                'listing': listing,
                'source': source,
            })

    # ---- 统计（包裹类型/发货SKU/三态/listing 分布基于平台信息表行；缺失行单独存放） ----
    stats = {
        'total': len(result),
        'original': len(rows),
        'expanded': n_added,
        'expanded_groups': len(expanded_groups),
        'pkg_dist': dict(Counter(r[fidx['包裹类型']] for r in result)),
        'fahuo_dist': dict(Counter(r[fidx['发货SKU']] for r in result)),
        'anomaly_count': sum(1 for r in result if r[fidx['包裹类型']] == '异常'),
        'tag_dist': dict(Counter(r[fidx['是否重复维护']] for r in result)),
        'asin_hit': n_asin,
        'listing_dist': dict(Counter(r[fidx[COL_LISTING]] for r in result
                                     if r[fidx[COL_LISTING]] != '')),
        'missing_platform': {               # 数据源有、平台信息表无 的 SKU（异常清单）
            'lingxing': len(lx_miss),
            'walmart': len(wm_miss),
            'union': len(union_miss),
            'rows': len(miss_rows),
        },
        'missing_platform_rows': miss_rows, # 独立面板数据：sku/spu/store/asin/listing/source
        'transfer': {                       # 本次使用的转单表口径
            'sheets': per_sheet,
            'total_rows': len(all_rows),
            'sku_count': len(mem2rep),
            'group_count': len(groups),
        },
        'lingxing': lx_summary,             # 本次使用的领星清单口径（None=未上传）
        'walmart': wm_summary,              # 本次使用的 walmart 报表口径（None=未上传）
    }
    return new_header, result, stats


# ========================== 对外入口 ==========================
def run_pipeline(raw_csv_path, xlsx_path, lingxing_path=None, walmart_path=None):
    """原始 CSV + 转单.xlsx + 可选(领星清单, walmart报表) -> (header, rows, stats)"""
    header, rows = _stage1(raw_csv_path)
    header, rows, stats = _stage2(header, rows, xlsx_path,
                                  lingxing_path=lingxing_path,
                                  walmart_path=walmart_path)
    return header, rows, stats


if __name__ == '__main__':
    import sys
    raw = sys.argv[1] if len(sys.argv) > 1 else r"D:/谷歌下载/SKU平台信息套件成员货品体现结果362.csv"
    xlsx = sys.argv[2] if len(sys.argv) > 2 else r"D:/WorkBuddy AI 文件/SKU&SPU清洗/system/reference/转单.xlsx"
    print("转单表摘要:", summarize_transfer(xlsx))
    h, rows, st = run_pipeline(raw, xlsx)
    print("列:", h)
    print("统计:", st)
